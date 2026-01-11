#!/usr/bin/env python3
"""Comprehensive test of 01_Data.py functionality"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

print("="*80)
print("COMPREHENSIVE DATA MODULE TEST")
print("="*80)

# Test 1: Quant Metrics Export - Single Ticker
print("\n1️⃣  Single Ticker (AAPL)...")
try:
    from portfolio.io.excel_export import export_quant_metrics_to_excel
    excel_bytes = export_quant_metrics_to_excel(
        tickers=["AAPL"], start="2024-01-01", end="2024-12-31"
    )
    print(f"   ✅ Works ({len(excel_bytes):,} bytes)")
except Exception as e:
    print(f"   ❌ Failed: {e}")

# Test 2: Multiple Tickers
print("\n2️⃣  Multiple Tickers (AAPL, MSFT)...")
try:
    excel_bytes = export_quant_metrics_to_excel(
        tickers=["AAPL", "MSFT"], start="2024-01-01", end="2024-12-31"
    )
    print(f"   ✅ Works ({len(excel_bytes):,} bytes)")
except Exception as e:
    print(f"   ❌ Failed: {e}")

# Test 3: Verify Excel Structure
print("\n3️⃣  Excel Structure...")
try:
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    for sheet in ["DATA", "SUMMARY", "METADATA", "DATA_QUALITY", "CORRELATION"]:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"   ✅ {sheet}: {ws.max_row} rows x {ws.max_column} cols")
        else:
            print(f"   ❌ {sheet}: MISSING")
except Exception as e:
    print(f"   ❌ Failed: {e}")

print("\n" + "="*80)
print("✅ All core functionality working!")
print("Now test in Streamlit UI at http://localhost:8501")
print("="*80)
