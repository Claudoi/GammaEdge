"""
Production-Grade Excel Export for Quant Metrics

Exports quantitative metrics to audit-ready Excel workbook with 5 sheets:
- DATA: Time series (date, ticker, adj_close, ret_1d)
- SUMMARY: Period metrics per ticker (beta, alpha, sharpe, etc.)
- METADATA: Standards and definitions
- DATA_QUALITY: Coverage, gaps, warnings
- CORRELATION: Correlation matrix + sample sizes (if ≥2 tickers)
"""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Literal

import polars as pl
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from portfolio.features.quant_metrics import (
    RF_ANNUAL_DEFAULT,
    RF_DAILY_DEFAULT,
    TRADING_DAYS_PER_YEAR,
    _as_pl_date,
    calculate_beta_alpha,
    calculate_cagr,
    calculate_calmar,
    calculate_correlation_matrix,
    calculate_data_quality,
    calculate_max_drawdown,
    calculate_moments,
    calculate_returns,
    calculate_sharpe_ratio,
)
from portfolio.features.quant_transforms import (
    compute_log_returns,
    compute_relative_volume,
    compute_intraday_volatility,
)


def export_quant_metrics_to_excel(
    tickers: list[str],
    start: str,
    end: str,
    benchmark: str = "SPY",
    rf_annual: float = RF_ANNUAL_DEFAULT,
    vol_lookback: int = 20,
    vol_method: Literal["parkinson", "garman_klass"] = "parkinson",
    output_format: Literal["bytes", "path"] = "bytes",
    output_path: str | None = None,
) -> bytes | str:
    """
    Export quant metrics to audit-grade Excel workbook.

    Args:
        tickers: List of ticker symbols
        start: Start date (YYYY-MM-DD)
        end: End date (YYYY-MM-DD)
        benchmark: Benchmark ticker for beta calculation (default: SPY)
        rf_annual: Annual risk-free rate (default: 0.02)
        vol_lookback: Lookback period for relative volume SMA (default: 20)
        vol_method: Volatility estimator ('parkinson' or 'garman_klass', default: 'parkinson')
        output_format: 'bytes' or 'path'
        output_path: Path to save file (required if output_format='path')

    Returns:
        bytes or path depending on output_format

    Raises:
        ValueError: If no valid data found or output_path missing

    Example:
        >>> excel_bytes = export_quant_metrics_to_excel(
        ...     tickers=["AAPL", "MSFT"],
        ...     start="2020-01-01",
        ...     end="2024-12-31",
        ... )
    """
    if output_format == "path" and not output_path:
        raise ValueError("output_path required when output_format='path'")

    # 1. Download data
    all_tickers = list(set(tickers + [benchmark]))
    df_prices = _download_prices(all_tickers, start, end)

    if df_prices.is_empty():
        raise ValueError(
            f"No price data downloaded for tickers: {tickers}. "
            f"Please check ticker symbols and date range ({start} to {end})."
        )

    # Validate that we have data for the requested tickers (not just benchmark)
    ticker_columns = [f"adj_close_{t}" for t in tickers]
    missing_tickers = [t for t in tickers if f"adj_close_{t}" not in df_prices.columns]
    
    if missing_tickers:
        raise ValueError(
            f"No data found for ticker(s): {', '.join(missing_tickers)}. "
            f"Please verify ticker symbols are correct."
        )

    # 2. Calculate returns
    df_returns = calculate_returns(df_prices, price_col_prefix="adj_close_")

    # 3. Get benchmark data
    bench_prices = df_prices.select(["date", f"adj_close_{benchmark}"]).rename(
        {f"adj_close_{benchmark}": "adj_close"}
    )
    bench_returns = df_returns.select(["date", f"ret_{benchmark}"]).rename(
        {f"ret_{benchmark}": "ret"}
    )
    bench_dates = bench_returns["date"]

    # 4. Calculate metrics for each ticker
    summary_rows = []
    data_quality_rows = []

    for ticker in tickers:
        price_col = f"adj_close_{ticker}"
        ret_col = f"ret_{ticker}"

        if price_col not in df_prices.columns or ret_col not in df_returns.columns:
            continue

        ticker_prices = df_prices.select(["date", price_col]).rename({price_col: "adj_close"})
        ticker_returns = df_returns.select(["date", ret_col]).rename({ret_col: "ret"})
        ticker_dates = ticker_returns["date"]

        # Beta & Alpha
        beta_result = calculate_beta_alpha(
            returns=ticker_returns["ret"],
            benchmark_returns=bench_returns["ret"],
            dates=ticker_dates,
            benchmark_dates=bench_dates,
        )

        # Sharpe
        sharpe_result = calculate_sharpe_ratio(
            returns=ticker_returns["ret"],
            rf_annual=rf_annual,
        )

        # Max Drawdown
        mdd_result = calculate_max_drawdown(
            prices=ticker_prices["adj_close"],
            dates=ticker_dates,
        )

        # CAGR
        cagr_result = calculate_cagr(
            prices=ticker_prices["adj_close"],
            dates=ticker_dates,
        )

        # Calmar
        calmar = calculate_calmar(
            cagr=cagr_result["cagr"],
            mdd=mdd_result["max_drawdown"],
        )

        # Moments
        moments_result = calculate_moments(ticker_returns["ret"])

        # Data Quality
        dq_result = calculate_data_quality(
            dates=ticker_dates,
            benchmark_dates=bench_dates,
            ticker=ticker,
        )

        # Summary row
        summary_rows.append({
            "ticker": ticker,
            "beta": beta_result["beta"],
            "alpha_daily": beta_result["alpha_daily"],
            "alpha_annual": beta_result["alpha_annual"],
            "r_squared": beta_result["r_squared"],
            "sharpe_ratio": sharpe_result["sharpe_ratio"],
            "max_drawdown": mdd_result["max_drawdown"],
            "cagr": cagr_result["cagr"],
            "calmar_ratio": calmar,
            "skewness": moments_result["skewness"],
            "kurtosis": moments_result["kurtosis"],
            "n_obs": beta_result["n_obs"],
        })

        # Data quality row
        data_quality_rows.append(dq_result)

    # 5. Create DataFrames
    df_summary = pl.DataFrame(summary_rows)
    df_data_quality = pl.DataFrame(data_quality_rows)

    # 6. Prepare DATA sheet (long format)
    df_data_long = _prepare_data_sheet(df_prices, df_returns, tickers, vol_lookback, vol_method)

    # 7. Calculate correlation (if ≥2 tickers)
    corr_result = None
    if len(tickers) >= 2:
        corr_result = calculate_correlation_matrix(df_returns, return_col_prefix="ret_")

    # 8. Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Generate sheets
    _generate_data_sheet(wb, df_data_long)
    
    _generate_summary_sheet(wb, df_summary)
    
    _generate_metadata_sheet(wb, start, end, benchmark, rf_annual, tickers, vol_lookback, vol_method)
    
    _generate_data_quality_sheet(wb, df_data_quality)

    if corr_result and corr_result["correlation_matrix"] is not None:
        _generate_correlation_sheet(
            wb,
            corr_result["correlation_matrix"],
            corr_result["sample_sizes"],
        )
    else:
        pass  # No correlation sheet for single ticker

    # 9. Save or return bytes
    if output_format == "path":
        wb.save(output_path)
        return output_path
    else:
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


def _download_prices(tickers: list[str], start: str, end: str) -> pl.DataFrame:
    """Download OHLCV data for multiple tickers.
    
    Returns:
        DataFrame with columns: date, open_{ticker}, high_{ticker}, low_{ticker}, 
        close_{ticker}, adj_close_{ticker}, volume_{ticker}
    """
    import pandas as pd
    
    # Download data (set auto_adjust=False to avoid warning)
    data = yf.download(tickers, start=start, end=end, progress=False, group_by="ticker", auto_adjust=False)
    
    # Check if data is empty
    if data is None or len(data) == 0:
        return pl.DataFrame()

    # Handle single vs multiple tickers
    if len(tickers) == 1:
        # Single ticker - columns are simple: Open, High, Low, Close, Adj Close, Volume
        df = data.reset_index()
        
        # Extract OHLCV columns
        cols_to_extract = ["Date"]
        col_mapping = {"Date": "date"}
        
        for col_name, prefix in [
            ("Open", "open"),
            ("High", "high"),
            ("Low", "low"),
            ("Close", "close"),
            ("Adj Close", "adj_close"),
            ("Volume", "volume"),
        ]:
            if col_name in df.columns:
                cols_to_extract.append(col_name)
                col_mapping[col_name] = f"{prefix}_{tickers[0]}"
        
        if len(cols_to_extract) == 1:  # Only Date, no price data
            return pl.DataFrame()
        
        df = df[cols_to_extract]
        df.columns = [col_mapping[c] for c in cols_to_extract]
        return pl.from_pandas(df)
    else:
        # Multiple tickers - MultiIndex columns
        if isinstance(data.columns, pd.MultiIndex):
            dfs = []
            for ticker in tickers:
                try:
                    if ticker in data.columns.levels[0]:
                        df_ticker = data[ticker].reset_index()
                        
                        # Extract OHLCV columns
                        cols_to_extract = ["Date"]
                        col_mapping = {"Date": "date"}
                        
                        for col_name, prefix in [
                            ("Open", "open"),
                            ("High", "high"),
                            ("Low", "low"),
                            ("Close", "close"),
                            ("Adj Close", "adj_close"),
                            ("Volume", "volume"),
                        ]:
                            if col_name in df_ticker.columns:
                                cols_to_extract.append(col_name)
                                col_mapping[col_name] = f"{prefix}_{ticker}"
                        
                        if len(cols_to_extract) == 1:  # Only Date
                            continue
                        
                        df_ticker = df_ticker[cols_to_extract]
                        df_ticker.columns = [col_mapping[c] for c in cols_to_extract]
                        dfs.append(pl.from_pandas(df_ticker))
                except (KeyError, AttributeError) as e:
                    # logger.warning(f"Error processing data for ticker {ticker}: {e}. Skipping.")
                    continue

            if not dfs:
                return pl.DataFrame()

            # Concatenate all dataframes and ensure no duplicate date columns
            if len(dfs) == 1:
                return dfs[0]
            
            # Get all unique dates
            all_dates = pl.concat([df.select("date") for df in dfs]).unique().sort("date")
            
            # Join each ticker's data
            result = all_dates
            for df in dfs:
                # Get all columns except date
                cols_to_join = [c for c in df.columns if c != "date"]
                result = result.join(df, on="date", how="left")

            return result
        else:
            # Fallback for unexpected format
            return pl.DataFrame()


def _pivot_to_long_ohlcv(df_wide: pl.DataFrame, tickers: list[str]) -> pl.DataFrame:
    """Convert wide OHLCV DataFrame to long format.
    
    Args:
        df_wide: Wide DataFrame with columns like open_{ticker}, high_{ticker}, etc.
        tickers: List of ticker symbols
    
    Returns:
        Long DataFrame with columns: date, ticker, open, high, low, close, adj_close, volume
    """
    if df_wide.is_empty():
        return pl.DataFrame()
    
    rows = []
    for ticker in tickers:
        # Build column mapping
        cols_map = {
            "date": "date",
        }
        cols_to_select = ["date"]
        
        for prefix in ["open", "high", "low", "close", "adj_close", "volume"]:
            col_name = f"{prefix}_{ticker}"
            if col_name in df_wide.columns:
                cols_map[col_name] = prefix
                cols_to_select.append(col_name)
        
        # Skip if we don't have at least close or adj_close
        if f"close_{ticker}" not in df_wide.columns and f"adj_close_{ticker}" not in df_wide.columns:
            continue
        
        # Select and rename columns
        df_ticker = df_wide.select(cols_to_select)
        df_ticker = df_ticker.rename(cols_map)
        
        # Add ticker column
        df_ticker = df_ticker.with_columns(pl.lit(ticker).alias("ticker"))
        
        # Ensure we have 'close' column (required by quant_transforms)
        if "close" not in df_ticker.columns and "adj_close" in df_ticker.columns:
            df_ticker = df_ticker.with_columns(pl.col("adj_close").alias("close"))
        
        # CRITICAL FIX: Normalize all numeric columns to Float64 to avoid schema mismatch
        # yfinance can return volume as Int64 or Float64 depending on the ticker
        # This causes pl.concat to fail with "type Float64 is incompatible with expected type Int64"
        numeric_cols = ["open", "high", "low", "close", "adj_close", "volume"]
        cast_exprs = []
        for col in numeric_cols:
            if col in df_ticker.columns:
                cast_exprs.append(pl.col(col).cast(pl.Float64))
        
        if cast_exprs:
            df_ticker = df_ticker.with_columns(cast_exprs)
        
        rows.append(df_ticker)
    
    if not rows:
        return pl.DataFrame()
    
    # Concatenate all tickers (now with consistent schema)
    df_long = pl.concat(rows).sort(["ticker", "date"])
    
    return df_long

def _prepare_data_sheet(
    df_prices: pl.DataFrame,
    df_returns: pl.DataFrame,
    tickers: list[str],
    vol_lookback: int = 20,
    vol_method: Literal["parkinson", "garman_klass"] = "parkinson",
) -> pl.DataFrame:
    """Prepare DATA sheet in long format with quant metrics.
    
    Returns:
        DataFrame with columns: date, ticker, adj_close, ret_1d, log_ret, rel_volume, intraday_vol
    """
    # First, convert wide OHLCV to long format for quant_transforms functions
    df_ohlcv_long = _pivot_to_long_ohlcv(df_prices, tickers)
    
    # Calculate quant metrics using quant_transforms
    # Ensure vol_lookback is int (not float)
    lookback_int = int(vol_lookback)
    
    df_log_ret = compute_log_returns(df_ohlcv_long) if not df_ohlcv_long.is_empty() else pl.DataFrame()
    df_rel_vol = compute_relative_volume(df_ohlcv_long, lookback=lookback_int) if not df_ohlcv_long.is_empty() else pl.DataFrame()
    df_intra_vol = compute_intraday_volatility(df_ohlcv_long, method=vol_method) if not df_ohlcv_long.is_empty() else pl.DataFrame()
    
    # Build base data (date, ticker, adj_close, ret_1d)
    rows = []
    for ticker in tickers:
        price_col = f"adj_close_{ticker}"
        ret_col = f"ret_{ticker}"

        if price_col not in df_prices.columns:
            continue

        df_ticker = df_prices.select(["date", price_col]).rename({price_col: "adj_close"})

        if ret_col in df_returns.columns:
            df_ticker = df_ticker.join(
                df_returns.select(["date", ret_col]).rename({ret_col: "ret_1d"}),
                on="date",
                how="left",
            )
        else:
            df_ticker = df_ticker.with_columns(pl.lit(None).alias("ret_1d"))

        df_ticker = df_ticker.with_columns(pl.lit(ticker).alias("ticker"))
        rows.append(df_ticker)

    if not rows:
        return pl.DataFrame()

    df_base = pl.concat(rows).select(["date", "ticker", "adj_close", "ret_1d"]).sort(["ticker", "date"])
    
    # Join quant metrics
    if not df_log_ret.is_empty():
        df_base = df_base.join(df_log_ret, on=["date", "ticker"], how="left")
    else:
        df_base = df_base.with_columns(pl.lit(None).alias("log_ret"))
    
    if not df_rel_vol.is_empty():
        df_base = df_base.join(df_rel_vol, on=["date", "ticker"], how="left")
    else:
        df_base = df_base.with_columns(pl.lit(None).alias("rel_volume"))
    
    if not df_intra_vol.is_empty():
        df_base = df_base.join(df_intra_vol, on=["date", "ticker"], how="left")
    else:
        df_base = df_base.with_columns(pl.lit(None).alias("intraday_vol"))
    
    return df_base.select(["date", "ticker", "adj_close", "ret_1d", "log_ret", "rel_volume", "intraday_vol"]).sort(["ticker", "date"])


def _generate_data_sheet(wb: Workbook, df: pl.DataFrame) -> None:
    """Generate DATA sheet with time series."""
    ws = wb.create_sheet("DATA", 0)

    # Check if DataFrame is empty
    if df.is_empty():
        ws.append(["date", "ticker", "adj_close", "ret_1d", "log_ret", "rel_volume", "intraday_vol"])
        _style_header(ws, 1)
        ws.append(["No data available"])
        return

    # Header
    ws.append(["date", "ticker", "adj_close", "ret_1d", "log_ret", "rel_volume", "intraday_vol"])
    _style_header(ws, 1)

    # Data
    for row in df.iter_rows():
        ws.append(row)

    # Format
    ws.column_dimensions["A"].width = 12  # date
    ws.column_dimensions["B"].width = 10  # ticker
    ws.column_dimensions["C"].width = 12  # adj_close
    ws.column_dimensions["D"].width = 12  # ret_1d
    ws.column_dimensions["E"].width = 12  # log_ret
    ws.column_dimensions["F"].width = 12  # rel_volume
    ws.column_dimensions["G"].width = 12  # intraday_vol


def _generate_summary_sheet(wb: Workbook, df: pl.DataFrame) -> None:
    """Generate SUMMARY sheet with period metrics."""
    ws = wb.create_sheet("SUMMARY", 1)

    # Header
    headers = [
        "ticker",
        "beta",
        "alpha_daily",
        "alpha_annual",
        "r_squared",
        "sharpe_ratio",
        "max_drawdown",
        "cagr",
        "calmar_ratio",
        "skewness",
        "kurtosis",
        "n_obs",
    ]
    ws.append(headers)
    _style_header(ws, 1)

    # Data
    for row in df.iter_rows():
        ws.append(row)

    # Format
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15


def _generate_metadata_sheet(
    wb: Workbook,
    start: str,
    end: str,
    benchmark: str,
    rf_annual: float,
    tickers: list[str],
    vol_lookback: int = 20,
    vol_method: Literal["parkinson", "garman_klass"] = "parkinson",
) -> None:
    """Generate METADATA sheet with standards and definitions."""
    ws = wb.create_sheet("METADATA", 2)

    metadata = [
        ["# Data Source", ""],
        ["provider", "yahoo"],
        ["price_field", "adj_close"],
        ["timezone", "America/New_York"],
        ["calendar_proxy", benchmark],
        ["", ""],
        ["# Returns", ""],
        ["returns_definition", "adj_close_to_adj_close_simple"],
        ["returns_formula", "(adj_close_t - adj_close_{t-1}) / adj_close_{t-1}"],
        ["", ""],
        ["# Quant Metrics", ""],
        ["log_returns_formula", "ln(close_t / close_{t-1})"],
        ["rel_volume_formula", "volume_t / SMA(volume, lookback)"],
        ["rel_volume_lookback", vol_lookback],
        ["intraday_vol_method", vol_method],
        ["parkinson_formula", "sqrt((ln(H/L))^2 / (4*ln(2)))"],
        ["garman_klass_formula", "sqrt(0.5*(ln(H/L))^2 - (2*ln(2)-1)*(ln(C/O))^2)"],
        ["", ""],
        ["# Annualization", ""],
        ["trading_days_per_year", TRADING_DAYS_PER_YEAR],
        ["rf_annual", rf_annual],
        ["rf_daily_value", RF_DAILY_DEFAULT],
        ["rf_daily_formula", "(1 + rf_annual)**(1/252) - 1"],
        ["", ""],
        ["# Benchmark", ""],
        ["benchmark_ticker", benchmark],
        ["", ""],
        ["# Statistical Methods", ""],
        ["correlation_method", "pearson"],
        ["skewness_method", "fisher_pearson_adjusted"],
        ["kurtosis_type", "excess"],
        ["", ""],
        ["# MDD", ""],
        ["mdd_sign", "negative"],
        ["mdd_formula", "min((equity_t / peak_t) - 1)"],
        ["", ""],
        ["# Sharpe", ""],
        ["sharpe_std", "std(returns)"],
        ["sharpe_formula", "mean(r - rf) / std(r) * sqrt(252)"],
        ["", ""],
        ["# Sample Sizes", ""],
        ["min_obs_sharpe", 60],
        ["min_obs_beta", 30],
        ["min_obs_cagr", 252],
        ["", ""],
        ["# Export Info", ""],
        ["date_range", f"{start} to {end}"],
        ["tickers", ", ".join(tickers)],
        ["generated_at", datetime.now().isoformat()],
        ["gammaedge_version", "1.0.0"],
    ]

    for row in metadata:
        ws.append(row)

    # Style headers (rows with #)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=len(metadata)), start=1):
        if row[0].value and str(row[0].value).startswith("#"):
            row[0].font = Font(bold=True, size=12)
            row[0].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def _generate_data_quality_sheet(wb: Workbook, df: pl.DataFrame) -> None:
    """Generate DATA_QUALITY sheet."""
    ws = wb.create_sheet("DATA_QUALITY", 3)

    # Header
    headers = [
        "ticker",
        "first_date",
        "last_date",
        "n_obs",
        "expected_obs",
        "coverage_pct",
        "max_gap_days",
        "missing_blocks",
        "warnings",
    ]
    ws.append(headers)
    _style_header(ws, 1)

    # Data
    for row in df.iter_rows():
        # Convert warnings list to string
        row_list = list(row)
        if row_list[-1]:  # warnings column
            if isinstance(row_list[-1], list):
                row_list[-1] = "; ".join(row_list[-1]) if row_list[-1] else ""
        else:
            row_list[-1] = ""
        ws.append(row_list)

    # Format
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15


def _generate_correlation_sheet(
    wb: Workbook,
    corr_df: pl.DataFrame,
    sample_df: pl.DataFrame,
) -> None:
    """Generate CORRELATION sheet with matrix and sample sizes."""
    ws = wb.create_sheet("CORRELATION", 4)

    # Correlation matrix
    ws.append(["Correlation Matrix"])
    ws["A1"].font = Font(bold=True, size=12)

    for row in corr_df.iter_rows():
        ws.append(row)

    _style_header(ws, 2)

    # Sample sizes
    start_row = len(corr_df) + 4
    ws.cell(row=start_row, column=1, value="Sample Sizes")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)

    for row in sample_df.iter_rows():
        ws.append(row)

    _style_header(ws, start_row + 1)

    # Format
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 12


def _style_header(ws, row_num: int) -> None:
    """Apply header styling to a row."""
    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
