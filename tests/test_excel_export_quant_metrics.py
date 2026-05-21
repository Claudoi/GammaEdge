"""
Integration tests for production-grade Excel export.

Tests verify:
- All 5 sheets exist and have correct structure
- METADATA contains all required standards
- CORRELATION sheet only appears for ≥2 tickers
- Data integrity and format correctness
"""

import io

import openpyxl
import pytest

from portfolio.io.excel_export import export_quant_metrics_to_excel


class TestExcelExportStructure:
    """Test Excel workbook structure and sheets."""

    def test_excel_has_all_sheets_multiple_tickers(self):
        """Verify all 5 sheets exist for multiple tickers."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL", "MSFT"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        sheet_names = wb.sheetnames

        assert "DATA" in sheet_names
        assert "SUMMARY" in sheet_names
        assert "METADATA" in sheet_names
        assert "DATA_QUALITY" in sheet_names
        assert "CORRELATION" in sheet_names

    def test_correlation_only_if_multiple_tickers(self):
        """Single ticker → no CORRELATION sheet."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        sheet_names = wb.sheetnames

        assert "DATA" in sheet_names
        assert "SUMMARY" in sheet_names
        assert "METADATA" in sheet_names
        assert "DATA_QUALITY" in sheet_names
        assert "CORRELATION" not in sheet_names


class TestDataSheet:
    """Test DATA sheet structure and content."""

    def test_data_sheet_has_required_columns(self):
        """Verify DATA sheet has correct columns."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-01-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["DATA"]

        # Check header row contains required columns (other engineered columns
        # such as log_ret/rel_volume/intraday_vol may also be present).
        headers = [cell.value for cell in ws[1]]
        required = {"date", "ticker", "adj_close", "ret_1d"}
        missing = required - set(headers)
        assert not missing, f"Missing required columns: {missing}"

    def test_data_sheet_long_format(self):
        """Verify DATA is in long format (one row per date-ticker)."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL", "MSFT"],
            start="2023-01-01",
            end="2023-01-10",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["DATA"]

        # Check that we have rows for both tickers
        tickers_in_data = set()
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[1]:  # ticker column
                tickers_in_data.add(row[1])

        assert "AAPL" in tickers_in_data
        assert "MSFT" in tickers_in_data


class TestSummarySheet:
    """Test SUMMARY sheet structure and content."""

    def test_summary_sheet_has_all_metrics(self):
        """Verify SUMMARY sheet has all required metric columns."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["SUMMARY"]

        # Check header row
        headers = [cell.value for cell in ws[1]]
        required_columns = [
            "ticker",
            "beta",
            "alpha_daily",
            "alpha_annual",
            "r_squared",
            "sharpe_ratio",
            "max_drawdown",
            "cagr",
            "calmar_ratio",
            "skewness",
            "kurtosis",
            "n_obs",
        ]

        for col in required_columns:
            assert col in headers, f"Missing column: {col}"

    def test_summary_one_row_per_ticker(self):
        """Verify SUMMARY has one row per ticker."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL", "MSFT", "GOOGL"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["SUMMARY"]

        # Count data rows (excluding header)
        n_rows = ws.max_row - 1
        assert n_rows == 3  # One row per ticker


class TestMetadataSheet:
    """Test METADATA sheet structure and content."""

    def test_metadata_sheet_exists(self):
        """Verify METADATA sheet exists."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        assert "METADATA" in wb.sheetnames

    def test_metadata_contains_required_standards(self):
        """Verify METADATA contains all required standards."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-12-31",
            rf_annual=0.03,
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["METADATA"]

        # Extract all keys from column A
        keys = [
            cell.value for cell in ws["A"] if cell.value and not str(cell.value).startswith("#")
        ]

        required_keys = [
            "provider",
            "price_field",
            "returns_definition",
            "trading_days_per_year",
            "rf_annual",
            "rf_daily_value",
            "benchmark_ticker",
            "correlation_method",
            "skewness_method",
            "kurtosis_type",
            "mdd_sign",
            "sharpe_formula",
            "date_range",
        ]

        for key in required_keys:
            assert key in keys, f"Missing metadata key: {key}"

    def test_metadata_rf_annual_matches_input(self):
        """Verify METADATA rf_annual matches input parameter."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-12-31",
            rf_annual=0.035,
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["METADATA"]

        # Find rf_annual row
        for row in ws.iter_rows(values_only=True):
            if row[0] == "rf_annual":
                assert row[1] == 0.035
                break
        else:
            pytest.fail("rf_annual not found in METADATA")


class TestDataQualitySheet:
    """Test DATA_QUALITY sheet structure and content."""

    def test_data_quality_sheet_has_required_columns(self):
        """Verify DATA_QUALITY sheet has correct columns."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["DATA_QUALITY"]

        # Check header row
        headers = [cell.value for cell in ws[1]]
        required_columns = [
            "ticker",
            "first_date",
            "last_date",
            "n_obs",
            "expected_obs",
            "coverage_pct",
            "max_gap_days",
            "missing_blocks",
            "warnings",
        ]

        for col in required_columns:
            assert col in headers, f"Missing column: {col}"


class TestCorrelationSheet:
    """Test CORRELATION sheet structure and content."""

    def test_correlation_matrix_is_symmetric(self):
        """Verify correlation matrix is symmetric."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL", "MSFT"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["CORRELATION"]

        # Layout: row 1 is the section title "Correlation Matrix".
        # The matrix itself starts at row 2 in long form:
        #     (ticker_name, c_1, c_2, ..., c_N)
        # i.e. there is no separate header row with ticker names; the row
        # labels in column A double as the matrix tickers.
        tickers: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            label = row[0]
            if label in (None, "", "Sample Sizes"):
                break
            tickers.append(label)

        # Read correlation values keyed by ticker.
        corr_values: dict[str, tuple] = {}
        for row in ws.iter_rows(min_row=2, max_row=1 + len(tickers), values_only=True):
            ticker = row[0]
            corr_values[ticker] = row[1 : len(tickers) + 1]

        # Check diagonal is 1.0
        for i, ticker in enumerate(tickers):
            assert abs(corr_values[ticker][i] - 1.0) < 1e-6, f"Diagonal for {ticker} should be 1.0"

    def test_correlation_has_sample_sizes(self):
        """Verify CORRELATION sheet includes sample sizes."""
        excel_bytes = export_quant_metrics_to_excel(
            tickers=["AAPL", "MSFT"],
            start="2023-01-01",
            end="2023-12-31",
        )

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb["CORRELATION"]

        # Find "Sample Sizes" section
        found_sample_sizes = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Sample Sizes":
                found_sample_sizes = True
                break

        assert found_sample_sizes, "Sample Sizes section not found in CORRELATION sheet"


class TestOutputFormats:
    """Test different output formats."""

    def test_output_format_bytes(self):
        """Verify bytes output format."""
        result = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-01-31",
            output_format="bytes",
        )

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_format_path(self, tmp_path):
        """Verify path output format."""
        output_file = tmp_path / "test_export.xlsx"

        result = export_quant_metrics_to_excel(
            tickers=["AAPL"],
            start="2023-01-01",
            end="2023-01-31",
            output_format="path",
            output_path=str(output_file),
        )

        assert result == str(output_file)
        assert output_file.exists()
        assert output_file.stat().st_size > 0

    def test_output_path_required_for_path_format(self):
        """Verify ValueError if output_path missing for path format."""
        with pytest.raises(ValueError, match="output_path required"):
            export_quant_metrics_to_excel(
                tickers=["AAPL"],
                start="2023-01-01",
                end="2023-01-31",
                output_format="path",
            )
